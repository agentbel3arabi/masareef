"""Excel bank statement parser using openpyxl."""

import datetime
import io

import openpyxl

from app.schemas.import_ import ParsedRow
from app.services.import_.row_validator import _DATE_FORMAT_MAP, validate_row


def get_sheet_names(raw_bytes: bytes) -> list[str]:
    """Return sheet names from an XLSX/XLS file."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    return list(wb.sheetnames)


def get_headers(raw_bytes: bytes, sheet_name: str | None = None, skip_rows: int = 0) -> list[str]:
    """Return column headers (first data row after skip_rows) using lazy iteration."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    # Consume and discard the first skip_rows rows
    for _ in range(skip_rows):
        try:
            next(rows_iter)
        except StopIteration:
            return []
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    return [str(c) if c is not None else "" for c in header_row]


def parse_excel(
    raw_bytes: bytes,
    column_mapping: dict[str, str],
    sheet_name: str | None = None,
    skip_rows: int = 0,
    date_format: str = "DD/MM/YYYY",
    currency: str = "EGP",
    currency_exponent: int = 2,
) -> list[ParsedRow]:
    """Parse an Excel file into ParsedRow list using the provided column mapping.

    Rows are streamed lazily via iter_rows() to avoid loading the entire
    worksheet into memory at once.
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    # Consume and discard the first skip_rows rows
    for _ in range(skip_rows):
        try:
            next(rows_iter)
        except StopIteration:
            return []
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(c) if c is not None else "" for c in header_row]

    def col_idx(field: str) -> int:
        header = column_mapping.get(field, "")
        return headers.index(header) if header in headers else -1

    date_idx = col_idx("date")
    desc_idx = col_idx("description")
    debit_idx = col_idx("debit")
    credit_idx = col_idx("credit")
    amount_idx = col_idx("amount")

    _date_fmt = _DATE_FORMAT_MAP.get(date_format, date_format)

    def cell_str(row_data: tuple, idx: int) -> str:
        if idx < 0 or idx >= len(row_data):
            return ""
        val = row_data[idx]
        if val is None:
            return ""
        if isinstance(val, (datetime.datetime, datetime.date)):
            d = val.date() if isinstance(val, datetime.datetime) else val
            return d.strftime(_date_fmt)
        return str(val)

    rows: list[ParsedRow] = []
    for idx, row_data in enumerate(rows_iter):  # stream data rows — no list() needed
        date_raw = cell_str(row_data, date_idx)
        desc_raw = cell_str(row_data, desc_idx)

        if amount_idx >= 0:
            row = validate_row(
                idx,
                date_raw,
                desc_raw,
                cell_str(row_data, amount_idx),
                "",
                date_format,
                currency,
                currency_exponent,
                single_amount=True,  # preserve sign from signed amount column
            )
        else:
            row = validate_row(
                idx,
                date_raw,
                desc_raw,
                cell_str(row_data, debit_idx),
                cell_str(row_data, credit_idx),
                date_format,
                currency,
                currency_exponent,
            )
        rows.append(row)

    return rows
