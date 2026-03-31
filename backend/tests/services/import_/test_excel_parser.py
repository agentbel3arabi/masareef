import io
import datetime
import pytest
import openpyxl

from app.services.import_.excel_parser import get_headers, get_sheet_names, parse_excel


def _make_xlsx(rows: list[list]) -> bytes:
    """Helper: create in-memory XLSX from a list of rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_BYTES = _make_xlsx([
    ["Date", "Description", "Debit", "Credit"],
    ["15/03/2026", "CARREFOUR", 1250.00, None],
    ["16/03/2026", "SALARY", None, 50000.00],
])

_MAPPING = {"date": "Date", "description": "Description", "debit": "Debit", "credit": "Credit"}


def test_get_headers():
    headers = get_headers(_XLSX_BYTES)
    assert headers == ["Date", "Description", "Debit", "Credit"]


def test_get_sheet_names():
    sheets = get_sheet_names(_XLSX_BYTES)
    assert len(sheets) >= 1


def test_parse_debit_row():
    rows = parse_excel(_XLSX_BYTES, _MAPPING, date_format="DD/MM/YYYY", currency="EGP")
    assert rows[0].status == "valid"
    assert rows[0].amount_minor == -125000


def test_parse_credit_row():
    rows = parse_excel(_XLSX_BYTES, _MAPPING, date_format="DD/MM/YYYY", currency="EGP")
    assert rows[1].amount_minor == 5000000
    assert rows[1].type == "credit"


def test_datetime_cell_parsed():
    """Excel may return datetime objects for date cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "Debit"])
    ws.append([datetime.datetime(2026, 3, 15), "MERCHANT", 100])
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    rows = parse_excel(raw, {"date": "Date", "description": "Description", "debit": "Debit"},
                       date_format="DD/MM/YYYY", currency="EGP")
    assert rows[0].date == datetime.date(2026, 3, 15)


def test_multi_sheet_selection():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Date", "Description", "Debit"])
    ws1.append(["15/03/2026", "MERCHANT", 100])
    ws2 = wb.create_sheet("Transactions")
    ws2.append(["Date", "Description", "Debit"])
    ws2.append(["16/03/2026", "SALARY", 200])
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    rows = parse_excel(raw, {"date": "Date", "description": "Description", "debit": "Debit"},
                       sheet_name="Transactions", date_format="DD/MM/YYYY", currency="EGP")
    assert len(rows) == 1
    assert rows[0].description == "SALARY"
